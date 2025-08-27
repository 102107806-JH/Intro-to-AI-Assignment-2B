from data_structures.graph_classes.adjacency_list_graph import AdjacencyListGraph
import re
import openpyxl

class GraphVertexEdgeInit:
    def __init__(self, graph_init_file_name):
        self._graph_init_file_name = graph_init_file_name

    def extract_file_contents(self):
        adjacency_list_graph = AdjacencyListGraph()
        adjacency_list_graph = self._insert_vertices(adjacency_list_graph)
        adjacency_list_graph = self._insert_edges(adjacency_list_graph)
        return adjacency_list_graph


    def _insert_vertices(self, adjacency_list_graph):
        workbook = openpyxl.load_workbook(self._graph_init_file_name)  # Open the file #
        sheet =  workbook.active  # Get the first sheet #

        for i in range(2, sheet.max_row + 1):
            # Extract all the appropriate info from each row and then add a vertex to the graph
            scats_cell = sheet.cell(row=i, column=1)
            long_cell = sheet.cell(row=i, column=4)
            lat_cell = sheet.cell(row=i, column=3)
            adjacency_list_graph.insert_vertex(int(scats_cell.value), long=float(long_cell.value), lat=float(lat_cell.value))

        return adjacency_list_graph

    def _insert_edges(self, adjacency_list_graph):
        workbook = openpyxl.load_workbook(self._graph_init_file_name)  # Open the file #
        sheet = workbook.active  # Get the first sheet #

        for i in range(2, sheet.max_row + 1):
            origin_scats_number = sheet.cell(row=i, column=1).value  # Extract the origin scats number #
            scats_neighbours_numbers = re.findall(r"\d+", sheet.cell(row=i, column=2).value)  # Gets the neighbour scats numbers as a list #

            for neighbour_num in scats_neighbours_numbers:  # Get the neighbour number #
                adjacency_list_graph.insert_edge_into_vertex(origin_scats_number=int(origin_scats_number), destination_scats_number=int(neighbour_num))  # Insert the as an edge to the origin node #

        return adjacency_list_graph





